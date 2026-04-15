"""
生成"清洗预览" Excel：完全从 parse_result.sheet1_data + mutations 重建一份全新工作簿，
**绝不重新打开用户原 Excel**——避开 openpyxl 在用户文件里碰到 vml drawing /
rich text 时的崩溃路径。

牺牲：丢掉用户原 Excel 的格式、Sheet 2、自定义公式等。换来的是稳定性 + 不用读原文件。
对"清洗预览"这个用途完全可以接受：用户要看的是哪些单元格被改了，不是要原样备份。
"""
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

HIGHLIGHT_FILL = PatternFill(start_color='FFFFCC00', end_color='FFFFCC00', fill_type='solid')  # 黄色：已修正
WARNING_FILL = PatternFill(start_color='FFFDE68A', end_color='FFFDE68A', fill_type='solid')   # 浅黄：需确认
REVERTED_FILL = PatternFill(start_color='FFCCCCCC', end_color='FFCCCCCC', fill_type='solid')   # 灰色：已撤回


# employee dict key → field_map key（两套命名不一致的地方）
_FIELD_ALIASES = {
    'base_annual': 'base_salary',
    'base_monthly': 'base_salary',
}


def _col_index(field: str, field_map: dict, column_names: list) -> int:
    """标准字段名 → Excel 列号（1-based）。找不到返回 -1。"""
    mapped = _FIELD_ALIASES.get(field, field)
    col_name = field_map.get(mapped)
    if col_name and col_name in column_names:
        return column_names.index(col_name) + 1
    return -1


def create_marked_excel(
    parse_result: dict,
    mutations: list,
    output_path: str,
    field_map: dict,
) -> str:
    """
    从 parse_result（含原始 sheet1_data + column_names）重建一份全新 xlsx，
    在上面叠加 mutation 标记。不读用户原 Excel。

    - reverted=True 的 mutation：跳过（保持原值，不染色）
    - 已修正（new_value 非空）：写入 new_value + 黄色 + 批注
    - 需确认（new_value=None）：保留原值 + 浅黄 + 批注
    """
    column_names = parse_result.get('column_names', [])
    sheet1_data = parse_result.get('sheet1_data', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # 写表头
    for c, name in enumerate(column_names, 1):
        ws.cell(row=1, column=c, value=name)

    # 写数据 —— sheet1_data 的 row_number 是 Excel 行号（2 起），row[i].data 是 {col_name: value}
    for entry in sheet1_data:
        row_num = entry.get('row_number')
        data = entry.get('data', {})
        if not row_num:
            continue
        for c, name in enumerate(column_names, 1):
            ws.cell(row=row_num, column=c, value=data.get(name))

    # 叠加 mutation 标记
    marked_count = 0
    for m in mutations:
        if m.get('reverted'):
            continue
        col_idx = _col_index(m['field'], field_map, column_names)
        if col_idx < 0:
            print(f'[ExcelMutator] field "{m["field"]}" not in field_map, skipping row {m.get("row_number")}')
            continue
        cell = ws.cell(row=m['row_number'], column=col_idx)
        old_val = cell.value

        if m.get('new_value') is not None:
            cell.value = m['new_value']
            cell.fill = HIGHLIGHT_FILL
            cell.comment = Comment(
                f"Sparky 已修正\n原始值: {old_val}\n{m.get('description', '')}",
                'Sparky'
            )
        else:
            cell.fill = WARNING_FILL
            cell.comment = Comment(
                f"Sparky 标记：需确认\n{m.get('description', '')}",
                'Sparky'
            )
        marked_count += 1

    print(f'[ExcelMutator] marked {marked_count}/{len(mutations)} cells in {output_path}')
    wb.save(output_path)
    wb.close()
    return output_path
