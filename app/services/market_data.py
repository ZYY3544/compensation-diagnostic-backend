"""
从 static/ 目录加载市场薪酬数据。启动时加载一次，缓存在内存中。

优先读 static/market_data.json（build 时由 scripts/build_market_data.py 预生成），
xlsx 仅作 fallback。这样运行时彻底脱离 openpyxl 的慢路径和 vml 崩溃风险。
xlsx 改了之后记得重跑 scripts/build_market_data.py 并 commit 新 JSON。
"""
import json
import os

_cache = {}


def _get_static_path(filename):
    return os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'static', filename)


def get_market_data():
    """加载市场薪酬数据，返回按 (job_function, hay_grade) 索引的字典"""
    if 'market_index' in _cache:
        return _cache['market_index']

    import time
    t0 = time.monotonic()

    json_path = _get_static_path('market_data.json')
    if os.path.exists(json_path):
        index = _load_from_json(json_path)
        _cache['market_index'] = index
        print(f'[market_data] loaded {len(index)} rows from JSON in {time.monotonic() - t0:.3f}s')
        return index

    # Fallback：JSON 不在（开发环境忘了 build），临时从 xlsx 现读
    xlsx_path = _get_static_path('市场薪酬数据.xlsx')
    if not os.path.exists(xlsx_path):
        _cache['market_index'] = {}
        return {}

    print('[market_data] WARN: market_data.json missing, falling back to xlsx (slow)')
    index = _load_from_xlsx(xlsx_path)
    _cache['market_index'] = index
    print(f'[market_data] loaded {len(index)} rows from xlsx in {time.monotonic() - t0:.2f}s')
    return index


def _load_from_json(path: str) -> dict:
    with open(path, 'r', encoding='utf-8') as f:
        records = json.load(f)
    return {(r['job_function'], int(r['hay_grade'])): r for r in records}


def _load_from_xlsx(path: str) -> dict:
    """xlsx fallback —— 用 iter_rows 单遍流式读取，O(N) 复杂度"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        first = next(rows_iter)
    except StopIteration:
        wb.close()
        return {}
    headers = [str(h).strip() if h is not None else '' for h in first]

    index = {}
    for row_values in rows_iter:
        row = {h: v for h, v in zip(headers, row_values) if h}
        job_func = str(row.get('Job Function', '')).strip()
        hay = row.get('Hay职级')
        if not job_func or hay is None:
            continue
        index[(job_func, int(hay))] = {
            'job_family': str(row.get('Job Family', '')).strip(),
            'job_function': job_func,
            'hay_grade': int(hay),
            'level': str(row.get('层级', '')).strip(),
            'base_p25': _safe_int(row.get('base_p25')),
            'base_p50': _safe_int(row.get('base_p50')),
            'base_p75': _safe_int(row.get('base_p75')),
            'bonus_p25': _safe_int(row.get('bonus_p25')),
            'bonus_p50': _safe_int(row.get('bonus_p50')),
            'bonus_p75': _safe_int(row.get('bonus_p75')),
            'ttc_p25': _safe_int(row.get('ttc_p25')),
            'ttc_p50': _safe_int(row.get('ttc_p50')),
            'ttc_p75': _safe_int(row.get('ttc_p75')),
        }
    wb.close()
    return index


def lookup_market_salary(job_function: str, hay_grade: int) -> dict | None:
    """查询特定职能+Hay职级的市场薪酬数据"""
    index = get_market_data()
    return index.get((job_function, hay_grade))


def get_all_job_functions() -> list[str]:
    """获取市场数据中所有 Job Function 名称"""
    index = get_market_data()
    return sorted(set(v['job_function'] for v in index.values()))


def _safe_int(val) -> int:
    if val is None:
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0
