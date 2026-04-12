import openpyxl


def parse_excel(file_path):
    """Parse uploaded Excel file"""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    result = {
        'sheet1_data': [],
        'sheet2_data': {},
        'column_names': [],
        'sheet_names': list(wb.sheetnames),
        'sheet_count': len(wb.sheetnames),
    }

    # Parse Sheet 1 (Employee data)
    if len(wb.sheetnames) >= 1:
        ws1 = wb[wb.sheetnames[0]]
        rows = list(ws1.iter_rows(values_only=True))
        if rows:
            result['column_names'] = [str(c) if c else f'Column_{i}' for i, c in enumerate(rows[0])]
            for i, row in enumerate(rows[1:], start=2):
                row_data = {}
                for j, val in enumerate(row):
                    if j < len(result['column_names']):
                        row_data[result['column_names'][j]] = val
                result['sheet1_data'].append({
                    'row_number': i,
                    'data': row_data
                })

    # Parse Sheet 2 (Company data) if exists
    if len(wb.sheetnames) >= 2:
        ws2 = wb[wb.sheetnames[1]]
        all_rows = list(ws2.iter_rows(values_only=True))

        # Detect year columns from header row
        years = []
        if all_rows:
            for cell in all_rows[0]:
                if cell:
                    s = str(cell).strip()
                    # Match "2022年" or "2022" or "FY2022"
                    import re
                    m = re.search(r'(20\d{2})', s)
                    if m:
                        years.append(int(m.group(1)))

        # Parse metrics: look for known keywords in column A
        metrics = []
        metric_keywords = {
            '营收': '年度营收',
            '收入': '年度营收',
            '利润': '年度利润',
            '员工总人数': '员工总人数',
            '人数': '员工总人数',
            '人工成本': '人工成本总额',
            '薪酬总额': '人工成本总额',
            '部门': '部门级绩效',
            '绩效': '部门级绩效',
        }
        seen_metrics = set()
        for row in all_rows[1:]:
            if not row or not row[0]:
                continue
            label = str(row[0]).strip()
            # Skip section headers like "一、经营数据"
            if label.startswith(('一', '二', '三', '四', '五')) and '、' in label:
                continue
            for kw, metric_name in metric_keywords.items():
                if kw in label and metric_name not in seen_metrics:
                    # Check if any year column has data
                    has_data = any(
                        row[j] is not None and str(row[j]).strip()
                        for j in range(2, min(len(row), 2 + len(years)))
                    ) if len(row) > 2 else False
                    metrics.append({
                        'name': metric_name,
                        'has_data': has_data,
                    })
                    seen_metrics.add(metric_name)
                    break

        # Also store as legacy key-value format
        for row in all_rows[1:]:
            if row and len(row) >= 2 and row[0]:
                result['sheet2_data'][str(row[0])] = row[1]

        result['sheet2_summary'] = {
            'years': sorted(years) if years else [],
            'year_count': len(years),
            'metrics': metrics,
        }
    else:
        result['sheet2_summary'] = {
            'years': [],
            'year_count': 0,
            'metrics': [],
        }

    wb.close()
    return result
